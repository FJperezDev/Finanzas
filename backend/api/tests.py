"""Tests del backend: validación, API, backup, exportación Excel y auth."""

import base64
import json
import shutil
import tempfile
from datetime import date
from io import BytesIO
from decimal import Decimal

from django.conf import settings
from django.core.cache import cache
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from PIL import Image

from .models import Backup, RefreshToken, Transaccion
from .seed import generar_filas_seed

from .models import (
    Backup,
    RefreshToken,
    Transaccion,
    Contacto,
    GastoCompartido,
    Participacion,
)


def tokens_de_admin() -> dict:
    """Inicia sesión con las credenciales admin del entorno y devuelve los tokens."""
    client = Client()
    respuesta = client.post(
        reverse("login"),
        data=json.dumps(
            {
                "username": settings.FINANZAS_ADMIN_USERNAME,
                "password": settings.FINANZAS_ADMIN_PASSWORD,
            }
        ),
        content_type="application/json",
    )
    assert respuesta.status_code == 200, respuesta.content
    return respuesta.json()


class AuthTests(TestCase):
    def setUp(self) -> None:
        # El throttle usa la caché compartida entre tests.
        cache.clear()

    def test_login_emite_access_y_refresh(self) -> None:
        tokens = tokens_de_admin()
        self.assertIn("access", tokens)
        self.assertIn("refresh", tokens)
        self.assertEqual(tokens["usuario"], settings.FINANZAS_ADMIN_USERNAME)
        self.assertEqual(RefreshToken.objects.filter(revocado=False).count(), 1)

    def test_login_con_credenciales_invalidas(self) -> None:
        respuesta = self.client.post(
            reverse("login"),
            data=json.dumps({"username": "admin", "password": "incorrecta"}),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 401)
        self.assertIn("errores", respuesta.json())

    def test_login_con_cuerpo_invalido(self) -> None:
        respuesta = self.client.post(
            reverse("login"),
            data="no es json",
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 400)

    def test_login_bloquea_tras_intentos_fallidos(self) -> None:
        for _ in range(5):
            respuesta = self.client.post(
                reverse("login"),
                data=json.dumps({"username": "admin", "password": "mala"}),
                content_type="application/json",
            )
            self.assertEqual(respuesta.status_code, 401)

        # El sexto intento queda bloqueado aunque las credenciales sean buenas.
        respuesta = self.client.post(
            reverse("login"),
            data=json.dumps(
                {
                    "username": settings.FINANZAS_ADMIN_USERNAME,
                    "password": settings.FINANZAS_ADMIN_PASSWORD,
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 429)

    def test_login_exitoso_resetea_el_contador(self) -> None:
        for _ in range(3):
            respuesta = self.client.post(
                reverse("login"),
                data=json.dumps({"username": "admin", "password": "mala"}),
                content_type="application/json",
            )
            self.assertEqual(respuesta.status_code, 401)

        tokens = tokens_de_admin()
        self.assertIn("access", tokens)

    def test_refrescar_rota_el_token(self) -> None:
        tokens = tokens_de_admin()
        refrescado = self.client.post(
            reverse("refrescar_token"),
            data=json.dumps({"refresh": tokens["refresh"]}),
            content_type="application/json",
        )
        self.assertEqual(refrescado.status_code, 200)
        self.assertNotEqual(refrescado.json()["refresh"], tokens["refresh"])

        # El refresh original queda revocado: reutilizarlo falla.
        reusado = self.client.post(
            reverse("refrescar_token"),
            data=json.dumps({"refresh": tokens["refresh"]}),
            content_type="application/json",
        )
        self.assertEqual(reusado.status_code, 401)

    def test_cerrar_sesion_revoca_el_refresh(self) -> None:
        tokens = tokens_de_admin()
        respuesta = self.client.post(
            reverse("cerrar_sesion"),
            data=json.dumps({"refresh": tokens["refresh"]}),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertTrue(RefreshToken.objects.get().revocado)

        refrescado = self.client.post(
            reverse("refrescar_token"),
            data=json.dumps({"refresh": tokens["refresh"]}),
            content_type="application/json",
        )
        self.assertEqual(refrescado.status_code, 401)

    def test_endpoint_protegido_sin_token(self) -> None:
        respuesta = self.client.get(reverse("listar_transacciones"))
        self.assertEqual(respuesta.status_code, 401)
        self.assertIn("errores", respuesta.json())

    def test_endpoint_protegido_con_token_malformado(self) -> None:
        respuesta = self.client.get(
            reverse("listar_transacciones"),
            HTTP_AUTHORIZATION="Bearer token-que-no-es-valido",
        )
        self.assertEqual(respuesta.status_code, 401)


class ValidacionTests(TestCase):
    def test_semilla_es_valida(self) -> None:
        from .validation import validar_filas

        errores = validar_filas(generar_filas_seed())
        self.assertEqual(errores, [])

    def test_detecta_todos_los_errores(self) -> None:
        from .validation import validar_filas

        filas = [
            {
                "Fecha": "2026-13-99",
                "Tipo": "X",
                "Categoria_Macro": "Y",
                "Subcategoria": "",
                "Concepto": "",
                "Importe": -5,
            },
            {
                "Fecha": "2026-01-01",
                "Tipo": "Ingreso",
                "Categoria_Macro": "Nómina",
                "Importe": "abc",
            },
        ]
        errores = validar_filas(filas)
        self.assertEqual(
            len(errores), 5
        )  # fecha, tipo, categoría, importe negativo y no numérico

    def test_columna_reservada_rechazada(self) -> None:
        from .validation import validar_filas

        filas = [
            {
                "Fecha": "2026-01-01",
                "Tipo": "Ingreso",
                "Categoria_Macro": "Nómina",
                "Importe": 100,
                "Periodo": "x",
            }
        ]
        errores = validar_filas(filas)
        self.assertTrue(any("reservado" in e for e in errores))


class ApiTests(TestCase):
    def setUp(self) -> None:
        cache.clear()
        Transaccion.objects.create(
            fecha=date(2026, 1, 1),
            tipo="Ingreso",
            categoria_macro="Nómina",
            subcategoria="Nomina",
            concepto="Nómina Enero",
            cuenta="Unicaja",
            importe=250000,
        )
        Transaccion.objects.create(
            fecha=date(2026, 1, 15),
            tipo="Gasto",
            categoria_macro="Fijo",
            subcategoria="Alquiler",
            concepto="Alquiler Piso",
            importe=115000,
            extras={"Bizum": "Sí"},
        )
        self.client.defaults["HTTP_AUTHORIZATION"] = (
            f"Bearer {tokens_de_admin()['access']}"
        )

    def test_listar_devuelve_filas_y_extras(self) -> None:
        respuesta = self.client.get(reverse("listar_transacciones"))
        self.assertEqual(respuesta.status_code, 200)
        datos = respuesta.json()
        self.assertEqual(len(datos["filas"]), 2)
        self.assertEqual(datos["filas"][0]["Cuenta"], "Unicaja")
        self.assertIn("Bizum", datos["columnas_extra"])
        self.assertEqual(datos["filas"][0]["Importe"], 2500.0)

    def test_guardar_valida_y_rechaza(self) -> None:
        malas = [
            {
                "Fecha": "2026-01-01",
                "Tipo": "Mal",
                "Categoria_Macro": "Nómina",
                "Importe": 10,
            }
        ]
        respuesta = self.client.post(
            reverse("guardar_transacciones"),
            data='{"filas": ' + str(malas).replace("'", '"') + "}",
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 400)
        self.assertIn("errores", respuesta.json())

    def test_guardar_hace_backup_y_reemplaza(self) -> None:
        filas = generar_filas_seed()
        respuesta = self.client.post(
            reverse("guardar_transacciones"),
            data='{"filas": ' + str(filas).replace("'", '"') + "}",
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.json()["filas"], len(filas))
        self.assertEqual(Transaccion.objects.count(), len(filas))
        self.assertEqual(Backup.objects.count(), 1)
        # El backup contiene el dataset previo (2 filas)
        self.assertEqual(len(Backup.objects.first().filas), 2)

    def test_guardar_con_columna_extra_conserva_extras(self) -> None:
        filas = [
            {
                "Fecha": "2026-01-01",
                "Tipo": "Ingreso",
                "Categoria_Macro": "Nómina",
                "Subcategoria": "Nómina",
                "Concepto": "Nómina",
                "Cuenta": "Revolut",
                "Importe": 2500,
                "Bizum": "Sí",
            }
        ]
        respuesta = self.client.post(
            reverse("guardar_transacciones"),
            data='{"filas": ' + str(filas).replace("'", '"') + "}",
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(Transaccion.objects.first().cuenta, "Revolut")
        self.assertEqual(Transaccion.objects.first().extras, {"Bizum": "Sí"})

    def test_exportar_descarga_xlsx(self) -> None:
        respuesta = self.client.get(reverse("exportar_transacciones"))
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Disposition"],
            'attachment; filename="transacciones.xlsx"',
        )
        self.assertEqual(respuesta.content[:2], b"PK")  # firma de archivo ZIP/xlsx

    def test_exportar_filtra_por_anio_y_mes(self) -> None:
        respuesta = self.client.get(
            reverse("exportar_transacciones"), {"anio": "2026", "mes": "1"}
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Disposition"],
            'attachment; filename="transacciones_2026_01.xlsx"',
        )
        hoja = load_workbook(BytesIO(respuesta.content)).active
        self.assertEqual(hoja.max_row, 3)  # cabecera + 2 filas de enero de 2026

    def test_exportar_filtra_por_anio_solo(self) -> None:
        respuesta = self.client.get(reverse("exportar_transacciones"), {"anio": "2025"})
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(
            respuesta["Content-Disposition"],
            'attachment; filename="transacciones_2025.xlsx"',
        )
        hoja = load_workbook(BytesIO(respuesta.content)).active
        self.assertEqual(hoja.max_row, 1)  # solo cabecera: no hay filas de 2025

    def test_exportar_rechaza_filtros_invalidos(self) -> None:
        respuesta = self.client.get(reverse("exportar_transacciones"), {"anio": "abc"})
        self.assertEqual(respuesta.status_code, 400)
        respuesta = self.client.get(reverse("exportar_transacciones"), {"mes": "13"})
        self.assertEqual(respuesta.status_code, 400)

    def test_health(self) -> None:
        respuesta = self.client.get(reverse("health"))
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.json()["transacciones"], 2)


class GastosCompartidosTests(TestCase):
    def setUp(self) -> None:
        cache.clear()
        self.client.defaults["HTTP_AUTHORIZATION"] = (
            f"Bearer {tokens_de_admin()['access']}"
        )

        # Crear contactos de prueba
        self.contacto_1 = Contacto.objects.create(nombre="Ana", telefono="+34611222333")
        self.contacto_2 = Contacto.objects.create(
            nombre="Carlos", telefono="+34644555666"
        )

    def test_crear_gasto_partes_iguales_pagador_yo(self) -> None:
        payload = {
            "concepto": "Cena Hamburguesas",
            "fecha": "2026-08-29",
            "importe_total": 60.00,
            "categoria_macro": "Ocio",
            "subcategoria": "Restaurantes",
            "tipo_reparto": "IGUALES",
            "pagador_id": None,  # Pago yo
            "participantes": [
                {"contacto_id": self.contacto_1.id},
                {"contacto_id": self.contacto_2.id},
            ],
        }
        respuesta = self.client.post(
            reverse("crear_gasto_compartido"),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)

        # 1. Verificar Gasto Compartido
        self.assertEqual(GastoCompartido.objects.count(), 1)
        gasto = GastoCompartido.objects.first()

        # 2. Verificar que SÍ se creó la Transacción en el Excel porque pagué yo
        self.assertIsNotNone(gasto.transaccion)
        self.assertEqual(gasto.transaccion.importe, 6000)  # Se guardó en céntimos
        self.assertEqual(gasto.transaccion.concepto, "Cena Hamburguesas")

        # 3. Verificar Participaciones (60€ / 3 personas = 20€ por participante)
        self.assertEqual(Participacion.objects.count(), 2)
        for p in Participacion.objects.all():
            self.assertEqual(p.importe_debido, Decimal("20.00"))

    def test_crear_gasto_reparto_exacto_pagador_amigo(self) -> None:
        payload = {
            "concepto": "Regalo Boda",
            "fecha": "2026-08-29",
            "importe_total": 150.00,
            "categoria_macro": "Regalo",
            "tipo_reparto": "EXACTO",
            "pagador_id": self.contacto_1.id,  # Paga Ana
            "participantes": [
                {
                    "contacto_id": self.contacto_2.id,
                    "importe_exacto": 50.00,
                }  # Carlos debe 50€
            ],
        }
        respuesta = self.client.post(
            reverse("crear_gasto_compartido"),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)

        gasto = GastoCompartido.objects.first()

        # 1. NO se debe haber generado transacción en el Excel porque pagó un amigo
        self.assertIsNone(gasto.transaccion)

        # 2. Se registra exactamente la deuda enviada
        participacion_carlos = Participacion.objects.get(contacto=self.contacto_2)
        self.assertEqual(participacion_carlos.importe_debido, Decimal("50.00"))

    def test_crear_gasto_exacto_rechaza_si_supera_total(self) -> None:
        payload = {
            "concepto": "Fraude",
            "fecha": "2026-08-29",
            "importe_total": 100.00,
            "categoria_macro": "Ocio",
            "tipo_reparto": "EXACTO",
            "pagador_id": None,
            "participantes": [
                {"contacto_id": self.contacto_1.id, "importe_exacto": 80.00},
                {
                    "contacto_id": self.contacto_2.id,
                    "importe_exacto": 50.00,
                },  # 80 + 50 = 130 > 100
            ],
        }
        respuesta = self.client.post(
            reverse("crear_gasto_compartido"),
            data=json.dumps(payload),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 400)
        self.assertIn("supera el total", respuesta.json()["errores"][0])

        # Verificar que el rollback de la BD funcionó y no hay deudas huérfanas
        self.assertEqual(GastoCompartido.objects.count(), 0)
        self.assertEqual(Participacion.objects.count(), 0)


def _png_1x1_b64() -> str:
    """Genera un PNG 1x1 válido y lo devuelve como data URL base64."""
    buf = BytesIO()
    Image.new("RGB", (1, 1), (255, 0, 0)).save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()


class SaldarDeudasTests(TestCase):
    def setUp(self) -> None:
        cache.clear()
        self.client.defaults["HTTP_AUTHORIZATION"] = (
            f"Bearer {tokens_de_admin()['access']}"
        )
        self.ana = Contacto.objects.create(nombre="Ana", telefono="+34611222333")
        self.carlos = Contacto.objects.create(
            nombre="Carlos", telefono="+34644555666"
        )

    def test_saldar_cuando_te_deben_crea_ingreso(self) -> None:
        # Pagaste tú: Ana te debe 20 €.
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("60.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("20.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "registrar_transaccion": True}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        datos = respuesta.json()
        self.assertEqual(datos["importe"], 20.0)
        self.assertEqual(datos["tipo"], "Ingreso")

        self.assertTrue(Participacion.objects.get().saldado)
        tx = Transaccion.objects.get(concepto__startswith="Saldar cuentas")
        self.assertEqual(tx.tipo, "Ingreso")
        self.assertEqual(tx.importe, 2000)  # 20 € en céntimos

    def test_saldar_cuando_debes_crea_gasto(self) -> None:
        # Pagó Ana un total de 100; Carlos debe 40, tu parte inferida es 60.
        gasto = GastoCompartido.objects.create(
            concepto="Regalo",
            fecha=date(2026, 8, 20),
            importe_total=Decimal("100.00"),
            categoria_macro="Regalo",
            subcategoria="Amigos",
            tipo_reparto="EXACTO",
            pagador=self.ana,
        )
        Participacion.objects.create(
            gasto=gasto, contacto=self.carlos, importe_debido=Decimal("40.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "registrar_transaccion": True}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        datos = respuesta.json()
        self.assertEqual(datos["importe"], 60.0)
        self.assertEqual(datos["tipo"], "Gasto")

        gasto.refresh_from_db()
        self.assertTrue(gasto.mi_parte_saldada)
        tx = Transaccion.objects.get(concepto__startswith="Saldar cuentas")
        self.assertEqual(tx.tipo, "Gasto")
        self.assertEqual(tx.importe, 6000)

    def test_saldar_sin_transaccion_no_crea_registro(self) -> None:
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("30.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("15.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "registrar_transaccion": False}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(Transaccion.objects.count(), 0)
        self.assertTrue(Participacion.objects.get().saldado)

    def test_actualizar_participacion(self) -> None:
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("30.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        participacion = Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("15.00")
        )

        respuesta = self.client.post(
            reverse("actualizar_participacion", args=[participacion.id]),
            data=json.dumps({"saldado": True}),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        participacion.refresh_from_db()
        self.assertTrue(participacion.saldado)
        self.assertEqual(participacion.importe_saldado, Decimal("15.00"))

        # Volver a abrirla
        respuesta = self.client.post(
            reverse("actualizar_participacion", args=[participacion.id]),
            data=json.dumps({"saldado": False}),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        participacion.refresh_from_db()
        self.assertFalse(participacion.saldado)
        self.assertEqual(participacion.importe_saldado, Decimal("0.00"))

    def test_saldar_parcial_deja_pendiente(self) -> None:
        # Ana te debe 20 € por una cena; saldas solo 8 €.
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("60.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        participacion = Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("20.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "importe": 8, "registrar_transaccion": True}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.json()["importe"], 8.0)

        participacion.refresh_from_db()
        self.assertEqual(participacion.importe_saldado, Decimal("8.00"))
        self.assertFalse(participacion.saldado)  # aún debe 12 €
        tx = Transaccion.objects.get(concepto__startswith="Saldar cuentas")
        self.assertEqual(tx.importe, 800)  # 8 € en céntimos

    def test_saldar_perdonando_marca_perdonado(self) -> None:
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("30.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        participacion = Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("15.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "registrar_transaccion": False}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        self.assertTrue(respuesta.json()["perdonado"])

        participacion.refresh_from_db()
        self.assertTrue(participacion.saldado)
        self.assertTrue(participacion.perdonado)
        self.assertEqual(Transaccion.objects.count(), 0)

    def test_saldar_sobrepago_vuelca_el_saldo(self) -> None:
        # Ana te debe 20 €; te paga 50 € (30 € de más → pasas a deberle 30).
        gasto = GastoCompartido.objects.create(
            concepto="Cena",
            fecha=date(2026, 8, 15),
            importe_total=Decimal("60.00"),
            categoria_macro="Ocio",
            subcategoria="Restaurantes",
            tipo_reparto="IGUALES",
            pagador=None,
        )
        participacion = Participacion.objects.create(
            gasto=gasto, contacto=self.ana, importe_debido=Decimal("20.00")
        )

        respuesta = self.client.post(
            reverse("saldar_gasto_compartido"),
            data=json.dumps(
                {"contacto_id": self.ana.id, "importe": 50, "registrar_transaccion": True}
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        datos = respuesta.json()
        self.assertEqual(datos["importe"], 50.0)
        self.assertEqual(datos["exceso"], 30.0)
        self.assertEqual(datos["tipo"], "Ingreso")

        participacion.refresh_from_db()
        self.assertTrue(participacion.saldado)

        # Se creó un gasto donde Ana pagó 30 → ahora le debo 30 €.
        nuevo = GastoCompartido.objects.get(concepto__startswith="Saldo a favor")
        self.assertEqual(nuevo.pagador, self.ana)
        self.assertEqual(nuevo.importe_total, Decimal("30.00"))

        # La transacción espejo es un ingreso de 50 €.
        tx = Transaccion.objects.get(concepto__startswith="Saldar cuentas")
        self.assertEqual(tx.tipo, "Ingreso")
        self.assertEqual(tx.importe, 5000)


class ContactoAvatarTests(TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        super().setUpClass()
        cls._media_tmp = tempfile.mkdtemp(prefix="finanzas_media_")
        cls._override = override_settings(MEDIA_ROOT=cls._media_tmp)
        cls._override.enable()

    @classmethod
    def tearDownClass(cls) -> None:
        cls._override.disable()
        shutil.rmtree(cls._media_tmp, ignore_errors=True)
        super().tearDownClass()

    def setUp(self) -> None:
        cache.clear()
        self.client.defaults["HTTP_AUTHORIZATION"] = (
            f"Bearer {tokens_de_admin()['access']}"
        )

    def test_crear_contacto_con_avatar_valido(self) -> None:
        respuesta = self.client.post(
            reverse("crear_contacto"),
            data=json.dumps(
                {
                    "nombre": "Laura",
                    "telefono": "+34699988877",
                    "icono": _png_1x1_b64(),
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 200)
        contacto = Contacto.objects.get(nombre="Laura")
        self.assertTrue(contacto.icono.name.startswith("iconos_contactos/"))
        self.assertIn("/media/iconos_contactos/", respuesta.json()["contacto"]["icono"])

    def test_crear_contacto_rechaza_avatar_no_imagen(self) -> None:
        respuesta = self.client.post(
            reverse("crear_contacto"),
            data=json.dumps(
                {
                    "nombre": "Laura",
                    "telefono": "+34699988877",
                    "icono": base64.b64encode(b"esto no es una imagen").decode(),
                }
            ),
            content_type="application/json",
        )
        self.assertEqual(respuesta.status_code, 400)
        self.assertFalse(Contacto.objects.filter(nombre="Laura").exists())

    def test_listar_contactos_devuelve_icono_nulo_sin_avatar(self) -> None:
        Contacto.objects.create(nombre="Ana", telefono="+34611222333")
        respuesta = self.client.get(reverse("listar_contactos"))
        self.assertEqual(respuesta.status_code, 200)
        self.assertIsNone(respuesta.json()["contactos"][0]["icono"])
